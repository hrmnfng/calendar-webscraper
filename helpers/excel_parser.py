'''
Used parse different Excel or CSV files given a input file and a data format.
'''

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime 
import sys
from pprint import pprint

class ExcelHelper:

    @staticmethod
    def read_file(file_path:str):
        
        if '.csv' in file_path:
            df = pd.read_csv(file_path)
        elif '.xlsx' in file_path:
            df = pd.read_excel(file_path, keep_default_na=False, na_values=[''])
            
            workbook = load_workbook(file_path, read_only=True)
            sheet=workbook.active

            for row_index, row in df.iterrows():
                for col_index, cell_value in enumerate(row):
                    cell = sheet.cell(row=row_index + 1, column=col_index + 1)
                    if pd.isna(cell_value):
                        df.iat[row_index, col_index] = ExcelHelper.get_merged_value(sheet, cell)
        
        return df

    @staticmethod
    def get_merged_value(sheet, cell):
        if any(cell.coordinate in merged_range for merged_range in sheet.merged_cells):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return merged_range.start_cell.value
        return cell.value

    @staticmethod
    def extract_dates(raw_data, search_name):
        
        # identify how many columns there are / weeks there are:
        column_names = raw_data.columns.tolist()

        # print(f'"{raw_data.columns[5]}"') "TUEDAY"
        # print(f'"{raw_data.columns[12]}"') "TUESDAY "

        # print(raw_data.info())
        possible_days = []
        for item in column_names:
            if type(raw_data.loc[0, item]) == datetime:
                # print(f'{item} - This is a a valid columm')
                possible_days.append([item, raw_data.loc[0, item]])
            # else:
            #     print(f'\t[{item}] - Nope')

        print(f'There are [{len(possible_days)}] dates in this spreadsheet')
        
        # identify which row to look at based on name
        search_name
        mask = raw_data['Select Name'].str.contains(search_name, case=False, na=False)
        indices_with_partial_match = raw_data.index[mask].tolist()
        if (len(indices_with_partial_match) > 0) and (len(indices_with_partial_match) > 1) :
            sys.exit('There is more than one match, please provide a more specific match-value')
        else:
            index = indices_with_partial_match[0]

        # Verify that we're getting the right rows of data
        # n = Name
        # n+1 = Shift
        # n+2 = Normal
        # n+3 = Overtime
        # n+4 = On Call 
        n = 5
        event_type = ['Shift', 'Normal', 'Overtime', 'On Call']
        calendar_events = []

        for i in range(1, n):
            row = index + i
            
            if raw_data.iloc[row, 1] == event_type[(i-1)]:
                print(f'This is correct - proceed to loop | value = [{raw_data.iloc[row, 1]}]')
                
                if event_type[(i-1)] == 'Shift': # we only care about shift right now
                    schedule = []
                    assignment = '' # set this outside of the loop, as some assignments span across multiple cells - should use the last available one if possible
                    
                    for item in possible_days:
                        col = item[0]
                        date = item[1]

                        # update the assignment, if available
                        print(f'\nAssignment = [{raw_data.loc[(index), col]}] in row [{index}] of column [{col}/{date}] ')
                        if type(raw_data.loc[(index), col]) == str: # using index because its at the top of the data
                            assignment = raw_data.loc[(index), col]

                        # try to extract date time, otherwise just paste the raw text and create an all day event
                        print(f'\tThe "shift" value is [{raw_data.loc[(row), col]}] in row [{row}] of column [{col}/{date}] ')
                        raw_time = raw_data.loc[row, col]
                        start_time, end_time = '', ''

                        # split the raw string to get start and end time
                        try:
                            split_times = raw_time.split('-') 
                            start_time = datetime.strptime(split_times[0], '%H%M').time() 
                            end_time = datetime.strptime(split_times[1], '%H%M').time()

                        # if a time is not provided, make it literally span all day - NEEDS TO BE CAUGHT LATER DOWN THE LINE #TODO
                        except Exception as e:
                            print(f'\tOops something went wrong:\n\t> {e}')
                            start_time = datetime.strptime('0000', '%H%M').time()
                            end_time = datetime.strptime('2359', '%H%M').time()

                        # create event start and end datetimes
                        event_start = datetime.combine(date.date(), start_time)
                        event_end = datetime.combine(date.date(), end_time)

                            # check if, Lord forbid, it goes past 11:59PM and into the next day
                            #TODO

                        # check for event name before compiling event details
                        if assignment == '':
                            event_name = '<No Data Here>'
                        else:
                            event_name = assignment
                        
                        roster_event = {event_name: {'start':event_start, 'end':event_end}}
                        
                        schedule.append(roster_event)

            else:
                print(f'You done screwed the pooch | value = [{raw_data.iloc[row, 1]}]; looking for [{event_type[i]}]')

        print('-----')
        return schedule


# row duplicate row names are done with .1, .2 ... at the end
    
if __name__ == "__main__":
    raw_data = ExcelHelper.read_file('./Term-3-Intern-Resident-Rosters.xlsx')   
    roster_schedule = ExcelHelper.extract_dates(raw_data=raw_data, search_name='FU, Michele')    

    pprint(raw_data)